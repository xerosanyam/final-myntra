from openpyxl import load_workbook

def userdata(mobile):

    wb = load_workbook('userdata.xlsx')
    sheet = wb.get_sheet_by_name('Sheet1')

    # for i in range(1,1):
    #     phone = str(sheet.cell(row=i, column=1).value)
    #     print phone
    #     if phone == str(mobile):
    print "excel data"
    #return sheet.cell(row=1, column=2).value, sheet.cell(row=1, column=3).value
    return "CAAJHwKVZCbFwBACeZAT5AkpOs2q9dVq09IOTZAPVrP6lnbgVBb4C5zgrvkvGjw5bpATKOLNLsoTGAqAxMDDA1R1jfZC8bihWANCh2GZCsHj5aYQ7DS2OOZCCfqzmzhM7yHfLrVwlQeJESQdiuMtP2zsZBxhtBA3PqjiszZAxQ0gDqSI7KSm8tPQ6B0rQW9jAhS1tHqxKu3gBHZAqnanHkg386", "10207883263017119"

#print userdata(8439257665)
